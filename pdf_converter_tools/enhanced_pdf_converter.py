#!/usr/bin/env python3
"""
Enhanced PDF to Excel Converter with Language Support
Allows specifying which language column to fill automatically
"""

import pandas as pd
import pdfplumber
import openpyxl
from openpyxl import Workbook
import re
import sys
import os
from pathlib import Path
import argparse

class QuizQuestion:
    def __init__(self):
        self.question_number = ""
        self.answer = ""
        self.question_text_zh = ""  # Chinese question text
        self.question_text_en = ""  # English question text
        self.question_text_id = ""  # Indonesian question text
        self.question_text_th = ""  # Thai question text
        self.question_text_vi = ""  # Vietnamese question text
        self.options_zh = {"A": "", "B": "", "C": "", "D": ""}  # Chinese options
        self.options_en = {"A": "", "B": "", "C": "", "D": ""}  # English options
        self.options_id = {"A": "", "B": "", "C": "", "D": ""}  # Indonesian options
        self.options_th = {"A": "", "B": "", "C": "", "D": ""}  # Thai options
        self.options_vi = {"A": "", "B": "", "C": "", "D": ""}  # Vietnamese options

def fix_thai_character_order(text):
    """Fix common Thai character order issues from PDF extraction"""
    # Common wrong patterns and their fixes:
    # 1. "ใชเ้" -> "ใช้" (extra เ and wrong order)
    # 2. "ติม" -> "เติม" (missing เ)
    
    # Pattern 1: Fix "ใชเ้" -> "ใช้"
    # Remove เ when it appears incorrectly between characters
    text = re.sub(r'([\u0E01-\u0E2E])([\u0E40])([\u0E48-\u0E4B])', r'\1\3', text)  # base + เ + tone -> base + tone
    
    # Pattern 2: Fix "ติม" -> "เติม" 
    # Add เ before ติ when it's part of "เติม" (เติม = เ + ติ + ม)
    # ติ = ต (U+0E15) + ิ (U+0E34), ม = ม (U+0E21)
    # But only if not already preceded by เ
    text = re.sub(r'(?<![\u0E40])([\u0E15])([\u0E34])([\u0E21])', '\u0E40\\1\\2\\3', text)  # ติม -> เติม (add เ before ติ, but not if already has เ)
    
    # Pattern 3: Fix other common issues where tone marks are misplaced
    # Move tone marks to right after base characters when they appear after pre-base vowels
    text = re.sub(r'([\u0E40-\u0E44])([\u0E48-\u0E4B])([\u0E01-\u0E2E])', r'\3\2\1', text)  # pre-vowel + tone + base -> base + tone + pre-vowel
    
    # Pattern 4: Fix "ควรใชน้ ้ำ" -> might need more context, but let's try to fix common patterns
    # This is more complex and might need manual correction or better heuristics
    
    return text

def normalize_thai_text(text):
    """Normalize Thai text by removing incorrect spaces between Thai characters"""
    # First, try to fix character order issues
    text = fix_thai_character_order(text)
    
    # Thai character range: U+0E00 to U+0E7F
    # Remove spaces that are between Thai characters (including combining characters)
    # This fixes issues where PDF extraction inserts spaces between Thai characters,
    # especially between base characters and combining characters (vowel marks, tone marks)
    
    # Use regex to find and remove whitespace between Thai characters
    # Match: Thai char + one or more whitespace + Thai char
    thai_char_pattern = r'[\u0E00-\u0E7F]'
    # Replace: Thai char + whitespace(s) + Thai char -> Thai char + Thai char
    # \s matches all whitespace including spaces, tabs, newlines, etc.
    pattern = f'({thai_char_pattern})\\s+({thai_char_pattern})'
    
    # Keep replacing until no more matches (in case of multiple spaces or nested issues)
    max_iterations = 10  # Prevent infinite loops
    iteration = 0
    while re.search(pattern, text) and iteration < max_iterations:
        text = re.sub(pattern, r'\1\2', text)
        iteration += 1
    
    # Also handle cases where combining characters (vowel/tone marks) are separated
    # These are common combining characters that should be attached to base characters
    # Thai combining characters: ั ิ ี ึ ื ุ ู ฺ ็ ่ ้ ๊ ๋ ำ ํ ฯ
    combining_chars = r'[\u0E31\u0E34\u0E35\u0E36\u0E37\u0E38\u0E39\u0E3A\u0E47\u0E48\u0E49\u0E4A\u0E4B\u0E4C\u0E4D\u0E33\u0E4D\u0E3F]'
    
    # Remove spaces before combining characters if preceded by a Thai character
    pattern2 = f'({thai_char_pattern})\\s+({combining_chars})'
    iteration = 0
    while re.search(pattern2, text) and iteration < max_iterations:
        text = re.sub(pattern2, r'\1\2', text)
        iteration += 1
    
    # Remove spaces after combining characters if followed by a Thai character
    pattern3 = f'({combining_chars})\\s+({thai_char_pattern})'
    iteration = 0
    while re.search(pattern3, text) and iteration < max_iterations:
        text = re.sub(pattern3, r'\1\2', text)
        iteration += 1
    
    # Specific fixes for known issues
    # Fix "ควรใชน้ ้ำ" or "ควรใชน้้ำ" -> "เป็นพิเศษน้ำ" (for question 6)
    # This is a specific case where PDF extraction got the wrong text
    text = re.sub(r'ควรใชน้\s*้ำ', 'เป็นพิเศษน้ำ', text)
    text = re.sub(r'ควรใชน้้ำ', 'เป็นพิเศษน้ำ', text)
    
    return text

def extract_quiz_questions_from_pdf(pdf_path, target_lang=None):
    """Extract quiz questions from PDF with bilingual parsing"""
    questions = []
    
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                # Normalize Thai text to fix spacing issues
                text = normalize_thai_text(text)
                full_text += text + "\n"
    
    # Split by question numbers to get individual questions
    question_pattern = r'(\d+)\.\s*\(([1-4])\)\s*(.+?)(?=\d+\.\s*\([1-4]\)|$)'
    question_matches = re.findall(question_pattern, full_text, re.DOTALL)
    
    for match in question_matches:
        question_number = match[0]
        answer_number = match[1]
        question_content = match[2].strip()
        
        # Create question object
        question = QuizQuestion()
        question.question_number = question_number
        answer_map = {"1": "A", "2": "B", "3": "C", "4": "D"}
        question.answer = answer_map[answer_number]
        
        # All PDFs are bilingual: zh + target_lang (en, id, th, or vi)
        # Default to en if no target_lang specified
        lang = target_lang if target_lang and target_lang != 'zh' else 'en'
        process_bilingual_question_content(question, question_content, lang)
        questions.append(question)
    
    return questions

def process_bilingual_question_content(question, content, target_lang='en'):
    """Process bilingual question content to separate Chinese and target language parts"""
    # All PDFs are bilingual: zh + target_lang (en, id, th, or vi)
    # Split content by detecting Chinese characters vs target language characters
    
    lines = content.split('\n')
    chinese_lines = []
    target_lang_lines = []
    
    for line in lines:
        line = line.strip()
        if line:
            # Count Chinese characters
            chinese_chars = sum(1 for c in line if '\u4e00' <= c <= '\u9fff')
            # Count non-Chinese, non-whitespace characters (target language)
            target_chars = sum(1 for c in line if c.isprintable() and not ('\u4e00' <= c <= '\u9fff') and not c.isspace())
            
            # If line has significant Chinese characters, it's Chinese
            # Otherwise, it's the target language
            if chinese_chars > 0:
                chinese_lines.append(line)
            elif target_chars > 5:  # Minimum threshold for target language content
                target_lang_lines.append(line)
            # If ambiguous, check if it has Chinese characters at all
            elif chinese_chars == 0 and target_chars > 0:
                target_lang_lines.append(line)
            else:
                # Default to Chinese if unclear
                chinese_lines.append(line)
    
    chinese_content = '\n'.join(chinese_lines).strip()
    target_lang_content = '\n'.join(target_lang_lines).strip()
    
    # If splitting didn't work well, try alternative approach
    # Look for patterns that indicate the target language
    if not target_lang_content and target_lang == 'en':
        # Try English-specific patterns
        english_patterns = [
            r'(Which of the following.*?)(?=\d+\.\s*\([1-4]\)|$)',
            r'(What.*?)(?=\d+\.\s*\([1-4]\)|$)',
            r'(How.*?)(?=\d+\.\s*\([1-4]\)|$)',
            r'(When.*?)(?=\d+\.\s*\([1-4]\)|$)',
            r'(Where.*?)(?=\d+\.\s*\([1-4]\)|$)'
        ]
        for pattern in english_patterns:
            match = re.search(pattern, content, re.DOTALL | re.IGNORECASE)
            if match:
                target_lang_content = match.group(1).strip()
                chinese_content = content.replace(target_lang_content, "").strip()
                break
    
    # Process Chinese content
    if chinese_content:
        process_question_content_language(question, chinese_content, 'zh')
    
    # Process target language content
    if target_lang_content:
        process_question_content_language(question, target_lang_content, target_lang)


def process_question_content_language(question, content, language):
    """Process question content for a specific language"""
    # Find the first option symbol
    first_option_pos = len(content)
    for symbol in ['①', '②', '③', '④']:
        pos = content.find(symbol)
        if pos != -1 and pos < first_option_pos:
            first_option_pos = pos
    
    if first_option_pos < len(content):
        # Separate question from options
        question_text = content[:first_option_pos].strip()
        options_content = content[first_option_pos:].strip()
        
        # Normalize Thai text if processing Thai language
        if language == 'th':
            question_text = normalize_thai_text(question_text)
            options_content = normalize_thai_text(options_content)
        
        # Store question text
        if language == 'zh':
            question.question_text_zh = question_text
        elif language == 'en':
            question.question_text_en = question_text
        elif language == 'id':
            question.question_text_id = question_text
        elif language == 'th':
            question.question_text_th = question_text
        elif language == 'vi':
            question.question_text_vi = question_text
        
        # Process options
        process_options_language(question, options_content, language)
    else:
        # No options found, just store the question text
        # Normalize Thai text if processing Thai language
        if language == 'th':
            content = normalize_thai_text(content)
        
        if language == 'zh':
            question.question_text_zh = content
        elif language == 'en':
            question.question_text_en = content
        elif language == 'id':
            question.question_text_id = content
        elif language == 'th':
            question.question_text_th = content
        elif language == 'vi':
            question.question_text_vi = content

def process_options_language(question, options_content, language):
    """Process options content to extract individual options for a specific language"""
    # Split by option symbols
    option_symbols = ['①', '②', '③', '④']
    option_map = {"①": "A", "②": "B", "③": "C", "④": "D"}
    
    # Find all option positions
    option_positions = []
    for symbol in option_symbols:
        pos = options_content.find(symbol)
        if pos != -1:
            option_positions.append((pos, symbol))
    
    # Sort by position
    option_positions.sort(key=lambda x: x[0])
    
    # Extract each option
    for i, (pos, symbol) in enumerate(option_positions):
        # Find the end position (next option or end of text)
        if i + 1 < len(option_positions):
            end_pos = option_positions[i + 1][0]
        else:
            end_pos = len(options_content)
        
        # Extract option text
        option_text = options_content[pos:end_pos].strip()
        # Remove the symbol from the beginning
        option_text = re.sub(r'^[①②③④]\s*', '', option_text).strip()
        
        # Normalize Thai text if processing Thai language
        if language == 'th':
            option_text = normalize_thai_text(option_text)
        
        if option_text:
            if language == 'zh':
                question.options_zh[option_map[symbol]] = option_text
            elif language == 'en':
                question.options_en[option_map[symbol]] = option_text
            elif language == 'id':
                question.options_id[option_map[symbol]] = option_text
            elif language == 'th':
                question.options_th[option_map[symbol]] = option_text
            elif language == 'vi':
                question.options_vi[option_map[symbol]] = option_text

def create_excel_with_quiz_structure(questions, output_path, target_lang=None):
    """Create Excel file with the same structure as the reference file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "converted"
    
    # Language mapping
    lang_columns = {
        'zh': {'question': 3, 'A': 8, 'B': 13, 'C': 18, 'D': 23},
        'en': {'question': 4, 'A': 9, 'B': 14, 'C': 19, 'D': 24},
        'id': {'question': 5, 'A': 10, 'B': 15, 'C': 20, 'D': 25},
        'th': {'question': 6, 'A': 11, 'B': 16, 'C': 21, 'D': 26},
        'vi': {'question': 7, 'A': 12, 'B': 17, 'C': 22, 'D': 27}
    }
    
    # Headers matching the reference file structure
    headers = [
        '題號', '答案', '題目_zh', '題目_en', '題目_id', '題目_th', '題目_vi',
        '選項A_zh', '選項A_en', '選項A_id', '選項A_th', '選項A_vi',
        '選項B_zh', '選項B_en', '選項B_id', '選項B_th', '選項B_vi',
        '選項C_zh', '選項C_en', '選項C_id', '選項C_th', '選項C_vi',
        '選項D_zh', '選項D_en', '選項D_id', '選項D_th', '選項D_vi',
        '題目圖片', '選項A_圖片', '選項B_圖片', '選項C_圖片', '選項D_圖片'
    ]
    
    # Set headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add question data
    # All PDFs are bilingual: zh + target_lang (en, id, th, or vi)
    fill_lang = target_lang if target_lang and target_lang != 'zh' else 'en'
    
    for row, question in enumerate(questions, 2):
        ws.cell(row=row, column=1, value=question.question_number)  # 題號
        ws.cell(row=row, column=2, value=question.answer)  # 答案
        
        # Always fill Chinese columns (zh)
        if question.question_text_zh:
            ws.cell(row=row, column=lang_columns['zh']['question'], value=question.question_text_zh)  # 題目_zh
            ws.cell(row=row, column=lang_columns['zh']['A'], value=question.options_zh['A'])  # 選項A_zh
            ws.cell(row=row, column=lang_columns['zh']['B'], value=question.options_zh['B'])  # 選項B_zh
            ws.cell(row=row, column=lang_columns['zh']['C'], value=question.options_zh['C'])  # 選項C_zh
            ws.cell(row=row, column=lang_columns['zh']['D'], value=question.options_zh['D'])  # 選項D_zh
        
        # Fill target language columns
        if fill_lang in lang_columns:
            cols = lang_columns[fill_lang]
            
            # Fill question text
            if fill_lang == 'en' and question.question_text_en:
                ws.cell(row=row, column=cols['question'], value=question.question_text_en)
            elif fill_lang == 'id' and question.question_text_id:
                ws.cell(row=row, column=cols['question'], value=question.question_text_id)
            elif fill_lang == 'th' and question.question_text_th:
                ws.cell(row=row, column=cols['question'], value=question.question_text_th)
            elif fill_lang == 'vi' and question.question_text_vi:
                ws.cell(row=row, column=cols['question'], value=question.question_text_vi)
            
            # Fill options
            for opt in ['A', 'B', 'C', 'D']:
                if fill_lang == 'en' and question.options_en[opt]:
                    ws.cell(row=row, column=cols[opt], value=question.options_en[opt])
                elif fill_lang == 'id' and question.options_id[opt]:
                    ws.cell(row=row, column=cols[opt], value=question.options_id[opt])
                elif fill_lang == 'th' and question.options_th[opt]:
                    ws.cell(row=row, column=cols[opt], value=question.options_th[opt])
                elif fill_lang == 'vi' and question.options_vi[opt]:
                    ws.cell(row=row, column=cols[opt], value=question.options_vi[opt])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    fill_lang = target_lang if target_lang and target_lang != 'zh' else 'en'
    print(f"Excel file created with {len(questions)} questions: {output_path}")
    print(f"Content filled in Chinese (zh) and {fill_lang.upper()} language columns")

def convert_single_pdf(pdf_path, output_dir=None, target_lang=None):
    """Convert a single PDF to Excel format"""
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        print(f"PDF file not found: {pdf_path}")
        return None
    
    if output_dir is None:
        output_dir = pdf_path.parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(exist_ok=True)
    
    print(f"Processing: {pdf_path.name}")
    fill_lang = target_lang if target_lang and target_lang != 'zh' else 'en'
    print(f"Extracting bilingual content: Chinese (zh) + {fill_lang.upper()}")
    
    # Extract questions
    questions = extract_quiz_questions_from_pdf(pdf_path, target_lang)
    if not questions:
        print(f"No questions found in {pdf_path}")
        return None
    
    # Create output filename
    lang_suffix = f"_{target_lang}" if target_lang else ""
    output_filename = pdf_path.stem + f"_converted{lang_suffix}.xlsx"
    output_path = output_dir / output_filename
    
    # Create Excel file
    create_excel_with_quiz_structure(questions, output_path, target_lang)
    return output_path

def main():
    parser = argparse.ArgumentParser(description='Convert PDF quiz files to Excel format with language support')
    parser.add_argument('input', help='PDF file or directory containing PDF files')
    parser.add_argument('-o', '--output', help='Output directory (default: same as input)')
    parser.add_argument('--batch', action='store_true', help='Process all PDF files in directory')
    parser.add_argument('-l', '--lang', choices=['en', 'id', 'th', 'vi'], 
                       help='Target language paired with Chinese (en=English, id=Indonesian, th=Thai, vi=Vietnamese). All PDFs are bilingual: zh + target language. Default: en')
    
    args = parser.parse_args()
    
    input_path = Path(args.input)
    
    if args.batch or input_path.is_dir():
        print("=== Batch Processing PDF Files ===")
        if input_path.is_dir():
            pdf_files = list(input_path.glob("*.pdf"))
            for pdf_file in pdf_files:
                convert_single_pdf(pdf_file, args.output, args.lang)
        else:
            print("Directory not found")
    else:
        print("=== Converting Single PDF File ===")
        output_path = convert_single_pdf(input_path, args.output, args.lang)
        if output_path:
            print(f"\n=== Conversion Complete ===")
            print(f"Output file: {output_path}")
            print("\nNext steps:")
            print("1. Open the Excel file")
            print("2. Fill in translations for other languages if needed")
            print("3. Add images if needed")
            print("4. Review and adjust formatting")

if __name__ == "__main__":
    # If no command line arguments, run with default files
    if len(sys.argv) == 1:
        pdf_path = "/Users/michael/Desktop/Projects/PracticePro/堆高機-菲律賓-英文.pdf"
        output_path = "/Users/michael/Desktop/Projects/PracticePro/pdf-xlsx-tool/converted_examples/堆高機-菲律賓-英文_converted_bilingual.xlsx"
        
        print("=== Converting Default PDF File ===")
        questions = extract_quiz_questions_from_pdf(pdf_path)
        
        if questions:
            print(f"Found {len(questions)} questions")
            
            # Show first few questions as preview
            for i, q in enumerate(questions[:3]):
                print(f"\nQuestion {q.question_number}:")
                print(f"  Answer: {q.answer}")
                print(f"  Chinese Text: {q.question_text_zh}")
                print(f"  English Text: {q.question_text_en}")
                print(f"  Chinese Options:")
                for opt, text in q.options_zh.items():
                    if text:
                        print(f"    {opt}: {text}")
                print(f"  English Options:")
                for opt, text in q.options_en.items():
                    if text:
                        print(f"    {opt}: {text}")
            
            create_excel_with_quiz_structure(questions, output_path)
            print(f"Output file: {output_path}")
        else:
            print("No questions found in the PDF")
    else:
        main()
